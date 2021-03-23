import openpyxl
import re
from PyQt5 import QtWidgets
from openpyxl.styles import PatternFill

from gui import Ui_HR_calculation
import sys


class hr_calc(QtWidgets.QMainWindow):
    def __init__(self):
        self.ui = Ui_HR_calculation()
        super().__init__()
        self.ui.setupUi(self)
        self.filename = ""
        self.res = 0
        self.count_days = list()
        self.counter_workers = 0
        self.from_date = list()  # с даты
        self.for_date = list()  # до даты
        self.dni_s = list()  # с даты
        self.dni_do = list()  # до даты
        self.ui.open_file.clicked.connect(lambda x: self.showDialog())
        self.ui.start.clicked.connect(lambda x: self.start_calc())
        self.cell_for_write_res = int()

    def indices(self, lst, element):
        result = []
        offset = -1
        while True:
            try:
                offset = lst.index(element, offset + 1)
            except ValueError:
                return result
            result.append(offset)

    def showDialog(self):
        self.filename = QtWidgets.QFileDialog.getOpenFileName(self, 'Open file', '*.xlsx')[0]
        print(self.filename)
        # name_index_ = fname.rfind("/")
        # self.filename = fname
        # self.ui.label_7.setText(fname[name_index_ + 1:])

    def f1(self):
        self.wb = openpyxl.load_workbook(self.filename, data_only=True)
        self.ws = self.wb[self.wb.sheetnames[0]]
        self.cell_for_write_res = self.ws.max_row + 2

        for row in self.ws.iter_rows(min_row=3, ):
            yield [cell.value for cell in row if cell.value]

    def convert_elem_to_int(self, lst):
        for i in range(0, len(lst)):
            lst[i] = int(lst[i])
        return lst

    def start_calc(self):

        # print(self.count_days)
        for day in range(1, int(self.ui.day_z.text()) + 1):
            self.count_days.append(day)
        for item in self.f1():

            for k in item:
                if type(k) == int or type(k) == float:
                    continue
                print(type(k), k)
                self.counter_workers += 1
                if re.search("до\s\d\d\.\d\d\.\d\d", k):
                    self.for_date.append(k)
                elif re.search("з\s\d\d\.\d\d\.\d\d", k):
                    self.from_date.append(k)


        for item in self.from_date:
            # print(item[self.indices(item, ".")[-2] - 2:self.indices(item, ".")[-2]])
            self.dni_s.append(item[self.indices(item, ".")[-2] - 2:self.indices(item, ".")[-2]])
        for item in self.for_date:
            self.dni_do.append(item[self.indices(item, ".")[-2] - 2:self.indices(item, ".")[-2]])

        self.convert_elem_to_int(self.dni_s)
        self.convert_elem_to_int(self.dni_do)
        self.counter_workers -= len(self.dni_s)
        self.ui.progressBar.setMaximum(int(self.ui.day_z.text()))
        for day in self.count_days:
            self.ui.progressBar.setValue(day)
            if day in self.dni_do:
                self.res += self.counter_workers
                self.counter_workers -= 1
                continue
            if day in self.dni_s:
                self.counter_workers += 1
                self.res += self.counter_workers
                continue
            self.res += self.counter_workers

        print(round(self.res / int(self.ui.day_z.text()), 2))  # dont forget get data from line edit
        print(self.cell_for_write_res)
        self.ws.cell(row=self.cell_for_write_res,
                     column=2).value = f"Результат: {round(self.res / int(self.ui.day_z.text()), 2)}"
        self.ws.cell(row=self.cell_for_write_res, column=2).fill = PatternFill("solid", fgColor="DDDDDD")
        self.wb.save(self.filename)


def main():
    app = QtWidgets.QApplication(sys.argv)
    w = hr_calc()
    w.show()
    app.exec_()


if __name__ == '__main__':
    main()
